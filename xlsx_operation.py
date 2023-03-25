from datetime import date
from pandas import ExcelWriter
import pandas as pd
from openpyxl import load_workbook, workbook
from create_sheet import create_sheet
import xlsxwriter
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Font

file = 'journal.xlsx'
wb = load_workbook(filename=file)
ws = wb['07-2022']
red_font = Font(color='00FF0000', italic=True, bold=True, )
green_font = Font(color='0099CC00', italic=True, bold=True,)
zero_font = Font(color='0000CCFF', italic=True, bold=True,)
stock_font = Font(color='009999FF', italic=True, bold=True)

# yellow 00FFFF00
# blue 0000CCFF
# red 00FF0000
# green 0099CC00
# grey 00C0C0C0
# dark green 00003300
# dark red 00800000
for i in range(2, 32):
    # print(i)
    green_cell = ws.cell(row=i, column=6)
    red_cell = ws.cell(row=i, column=7)
    stock_cell = ws.cell(row=i, column=4)
    des_cell = ws.cell(row=i, column=8)
    print(green_cell.value)

    if red_cell.value and green_cell.value:
        stock_cell.font = stock_font
        print(green_cell.value, red_cell.value)
        green_cell.font = green_font
        red_cell.font = red_font
        if red_cell.value == 'None':
            red_cell.font = zero_font
        if green_cell.value == 'None':
            green_cell.font = zero_font
            # green_cell.fill = PatternFill(
            #     start_color="0000CCFF", end_color="0000CCFF", fill_type="solid")
    else:
        break

wb.save(filename=file)
# print(profit_df)
# loss_df = df[pd.notnull(df['Loss_Price'])]


# df = pd.DataFrame([
#     {"a": True},
#     {"a": False},
#     {"a": True},
#     {"a": False},
# #     {"a": False},
# # ])
# # df["a"] = df["a"].map({True: "True", False: "False"})
# with pd.ExcelWriter("journal.xlsx", engine="openpyxl") as writer:
#     sheet_name = "Bool"
#     # Export DataFrame content
#     df.to_excel(writer, sheet_name=sheet_name)
#     # Set backgrund colors depending on cell values
#     sheet = writer.sheets[sheet_name]
#     # Skip header row, process as many rows as there are DataFrames
#     for cell, in sheet[f'B2:B{len(df) + 1}']:
#         value = df["a"].iloc[cell.row - 2]  # value is "True" or "False"
#         cell.fill = PatternFill("solid", start_color=(
#             "5cb800" if value == "True" else 'ff2800'))
# writer = pd.ExcelWriter('journal.xlsx', engine='xlsxwriter')
# wb = pd.ExcelFile('journal.xlsx')
# sheets = wb.sheet_names

# # worksheet = writer.sheets['Sheet1']
# print(sheets)
# for i in worksheet:
#     i.set_column("A:A", 12)
#     cell_format_green = workbook.add_format()
#     cell_format_green.set_bg_color('green')
#     worksheet.conditional_format('H2:I{}'.format(last_row), {'type':     'formula',
#                                                              'criteria': '=$H2<$I2',
#                                                              'format':   cell_format_red})
#     writer.close()

# Remember to close the writer
# get_date = date.today()
# today_date = get_date.strftime('%Y-%m-%d')
# print(today_date.split('-')[2])

# for i in date_lst:
#     date_in_lst = i.split('-')[2]
#     month = i.split('-')[1]
#     month_today = today_date.split('-')[1]
#     if month == month_today:
#         if '20' in date_in_lst:
#             print('ok')
#         else:
#             print('np')
