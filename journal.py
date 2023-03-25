import os
import pandas as pd
from datetime import date
from pandas import ExcelWriter
from openpyxl import load_workbook, workbook
from create_sheet import create_sheet, style_sheet



def add_func(lst):
    count = 0
    for i in lst:
        if i == "None":
            i = 0
            count = count+int(i)
        else:
            count = count+int(i)
    return count


def get_overall(filename):
    wb = pd.ExcelFile(filename)
    sheets = wb.sheet_names
    print(len(sheets))
    for i in sheets:
        if '-' in i or len(sheets) >= 1:
            # print(today_date.split('-')[1], i.split('-')[0])
            if i.split('-')[0] == today_date.split('-')[1]:
                df = pd.read_excel(filename, sheet_name=i)
                profit_df = df[pd.notnull(df['Profit_Price'])]
                loss_df = df[pd.notnull(df['Loss_Price'])]
                plst = profit_df['Profit_Price'].tolist()
                llst = loss_df['Loss_Price'].tolist()
                overall_p = add_func(plst)
                # print(overall_p)
                overall_l = add_func(llst)
                # print(overall_l)
                return [overall_p, overall_l]

        else:
            print('something is wrong sheets or sheet name')


def write_journal(filename, index):
    profit = ''
    loss = ''
    stk_name = input('Enter the stock name:')
    P_L = input('Profit/Loss:')
    P_L = P_L.capitalize()
    if P_L == "P":
        profit = input('Enter the Profit price:')
        loss = 'None'
    else:
        loss = input('Enter the Loss price:')
        profit = "None"
    Des = input('Any points to noted:')

    if os.path.exists(filename):

        if today_date.split('-')[2] == '01':
            create_sheet(today_date, filename)
        overall = get_overall(filename)
        
        if get_day.day_name() == 'Friday':
            print('total profit and losses of the previous weeks : ',
                  overall[0], overall[1])
        if today_date.split('-')[2] == '30' or today_date.split('-')[2] == '31':
            # overall_profit, overall_loss = get_overall()
            print('The over all profits for this month were : ',
                  overall[0])
            print('The over all Losses for this month were : ',
                  overall[1])
            if overall[0] > overall[1]:
                print('this month got success full trades : total value is ->$$',
                      overall[0]-overall[1])
            else:
                print('this month many trades were failed :total lossses is ->$$',
                      overall[1]-overall[0])


        sheet_title = today_date.split('-')[1]+'-'+today_date.split('-')[0]
        print('File exists writing data into excel .....')
        wb = load_workbook(filename)
        ws = wb[sheet_title]
        ws.append([index, today_date, get_day.day_name(),
                  stk_name.capitalize(), P_L, profit, loss, Des])
        wb.save(filename)
    else:
        sheet_title = today_date.split('-')[1]+'-'+today_date.split('-')[0]

        print('File dosent exists ,creating excel file ........')
        data_dict = {'Date': '',
                     'Day': '', 'Stock_Name': '', 'Profit/Loss': '', 'Profit_Price': '', 'Loss_Price': '', 'Description': ''}
        data_dict['Date'] = today_date
        data_dict["Day"] = get_day.day_name()
        data_dict["Stock_Name"] = stk_name
        data_dict["Profit/Loss"] = P_L
        data_dict["Profit_Price"] = profit
        data_dict["Loss_Price"] = loss
        data_dict["Description"] = Des
        writer = ExcelWriter('journal.xlsx')
        df = pd.DataFrame(data_dict, index=[index])
        df.to_excel(writer, sheet_title, index=True)
        writer.save()
    
    styling = style_sheet(today_date, filename)
    print(f'Styling status  is {styling}')


if __name__ == '__main__':
    filename = 'journal.xlsx'
    get_date = date.today()
    today_date = get_date.strftime('%Y-%m-%d')
    get_day = pd.Timestamp(today_date)



    for i in range(int(input('No of trades taken :'))):
        write_journal(filename, i+1)

# get_overall()
