import pandas as pd
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


# get_date = date.today()
# today_date = get_date.strftime('%Y-%m-%d')


def create_sheet(today_date, filename):

    wb = pd.ExcelFile(filename)
    sheets = wb.sheet_names
    sheet_title = today_date.split('-')[1]+'-'+today_date.split('-')[0]

    if sheet_title in sheets:
        print('Sheet already exists')
    else:
        print(f'creating excel Sheet in {filename}')
        wb = load_workbook(filename)
        print(sheets)
        wb.create_sheet(sheet_title)
        wb.save(filename)


def style_sheet(today_date, file):
    sheet_title = today_date.split('-')[1]+'-'+today_date.split('-')[0]
    print(sheet_title)
    wb = load_workbook(filename=file)
    ws = wb[str(sheet_title)]
    red_font = Font(color='00FF0000', italic=True, bold=True, )
    green_font = Font(color='0099CC00', italic=True, bold=True,)
    zero_font = Font(color='0000CCFF', italic=True, bold=True,)
    stock_font = Font(color='009999FF', italic=True, bold=True)
    des_font = Font(color='00808000', italic=True, bold=True)
    # date_font = Font(color='00FF8080', italic=True, bold=True)
    day_font = Font(color='00FF6600', italic=True, bold=True)
    # yellow 00FFFF00
    # blue 0000CCFF
    # red 00FF0000
    # green 0099CC00
    # grey 00C0C0C0
    # dark green 00003300
    # dark red 00800000
    for i in range(2, 32):
        # print(i)
        green_cell = ws.cell(row=i, column=6)
        red_cell = ws.cell(row=i, column=7)
        stock_cell = ws.cell(row=i, column=4)
        p_l_cell = ws.cell(row=i, column=5)
        # date_cell = ws.cell(row=i, column=2)
        day_cell = ws.cell(row=i, column=3)
        des_cell = ws.cell(row=i, column=8)
        if red_cell.value and green_cell.value:
            if p_l_cell.value == 'P':
                p_l_cell.font = green_font
            else:
                p_l_cell.font = red_font
            stock_cell.font = stock_font
            des_cell.font = des_font
            day_cell.font = day_font
            # date_cellfont = date_font
            green_cell.font = green_font
            red_cell.font = red_font
            if red_cell.value == 'None':
                red_cell.font = zero_font
            if green_cell.value == 'None':
                green_cell.font = zero_font
                # green_cell.fill = PatternFill(
                #     start_color="0000CCFF", end_color="0000CCFF", fill_type="solid")
        else:
            break
    wb.save(filename=file)

    return True


# style_sheet(today_date, 'journal.xlsx')
